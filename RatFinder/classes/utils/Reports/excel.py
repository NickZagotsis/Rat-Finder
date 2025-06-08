import sys
import openpyxl as xl
from os import makedirs
from os.path import exists
from os.path import join as pjoin
from openpyxl.styles import Alignment, PatternFill

class Excel:
    """
    This class is used for the generation of excel reports of all RAT instances provided.
    Attributes
    --------
    rat : Anydesk | Teamviewer
        The RAT instance to be used for the generation of excel reports.
    out : str
        The output directory for the excel reports.
    """
    def __init__(self, rat):
        self.rat = rat
        self.out = pjoin(self.rat.shared.output, 'excel')
        if not exists(self.out):
            makedirs(self.out)

    @staticmethod
    def __setup_worksheet(sheet, columns, widths):
        """
        Static helper function that sets up the worksheet with headers and column widths.

        Arguments
        ---------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to set up.
            columns : list[str]
                The list of column headers.
            widths : list[float|int]
                The list of column widths.
        """
        # Set up worksheet headers and column widths
        # Set column headers
        for i, header in enumerate(columns):
            col = chr(65 + i)
            sheet[f'{col}1'] = header

        # Set column widths
        for i, width in enumerate(widths):
            col = chr(65 + i)
            sheet.column_dimensions[col].width = width

    @staticmethod
    def __write_trace(sheet, ad_result):
        """
        Static helper function that writes the trace data to the worksheet.
        Arguments
        ---------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to write the data to.
            ad_result : dict[int,list[dict[str, str]]]
                The dictionary containing the trace data.
        """
        i = 2
        for session_key, values in ad_result.items():
            for data in values:
                for key_id, key in enumerate(list(data.keys()),0):
                    try:
                        if key == "File":
                            cell_ref = f"G{i}"
                            sheet[f'H{i}'] = session_key
                        else:
                            cell_ref = f'{chr(65+key_id)}{i}'
                        sheet[cell_ref] = data[key]
                        sheet[cell_ref].alignment = Alignment(wrap_text=True)
                        if key == "Explanation":
                            sheet[f"F{i}"] = data[key]
                            if data[key]:
                                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                sheet[cell_ref].fill = red_fill
                                sheet[f"E{i}"].fill = red_fill

                    except KeyboardInterrupt:
                        sys.exit()
                    except Exception as e:
                        sheet[f'{chr(65 + key_id)}{i}'] = "Err"
                i+=1

    def __write_trace_sheet_helper(self, wb, sheet_name, result):
        """
        This function is used to call the write_trace function and set up the worksheets.

        Arguments
        ---------
        wb : openpyxl.Workbook
            The workbook to write the data to.
        sheet_name : str
            The name of the sheet to write the data to.
        result : dict[int,list[dict[str, str]]]
            The dictionary containing the trace data.
        """
        if result:
            sheet = wb[sheet_name]
            # Set up column headers and widths
            widths = [8, 26.7, 10, 15.14, 52, 56, 90, 8]
            columns = ["LogLevel", "Timestamp", "Service", "LogService", "Message", "Explanation", "File", "Session"]
            Excel.__setup_worksheet(sheet, columns, widths)
            self.__write_trace(sheet,result)

    @staticmethod
    def __write_connection_trace(sheet, connection_result):
        """
        Helper function that writes the connection trace data to the worksheet.

        Arguments
        ---------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to write the data to.
            connection_result : list[str]
                The list of connection trace data.
        """
        for i, data in enumerate(connection_result, 2):
            for j, toinsert in enumerate(data.split("#")):
                try:
                    if 65 + j > 69:
                        continue
                    cell_ref = f'{chr(65 + j)}{i}'
                    sheet[cell_ref] = str(toinsert)
                except Exception as e:
                    sheet[f'{chr(65 + j)}{i}'] = "Err"

    @staticmethod
    def write_file_trace(sheet, file_trace):
        for i, row in enumerate(file_trace, start=2):
            for j, data in enumerate(row.values()):
                try:
                    cell_ref = f'{chr(65 + j)}{i}'
                    sheet[cell_ref] = str(data)
                except Exception as e:
                    print(e)
                    sheet[f'{chr(65 + j)}{i}'] = "Er"

    def write_anydesk(self, ad_svc_result=None, ad_result=None, connection_result=None, file_trace=None):
        """
        Base function used to write the Anydesk results to an excel file.

        Arguments
        ---------
            ad_svc_result : dict[int,list[dict[str, str]]]
                The dictionary containing the ad_svc.trace data.
            ad_result : dict[int,list[dict[str, str]]]
                The dictionary containing the ad.trace data.
            connection_result : list[str]
                The list of connection trace data.
        """
        if not ad_svc_result:
            ad_svc_result = self.rat.ad_svc_trace_results
        if not ad_result:
            ad_result = self.rat.ad_trace_results
        if not connection_result:
            connection_result = self.rat.connection_trace_results
        if not file_trace:
            file_trace = self.rat.file_trace_results

        outdir = pjoin(self.rat.shared.output, 'excel')

        # Try to load existing workbook or create new one
        try:
            wb = xl.load_workbook(pjoin(outdir, "RatFinderResults_Anydesk.xlsx"))
        except FileNotFoundError:
            wb = xl.Workbook()
        except Exception as E:
            self.rat.shared.logger.bind(category="anydesk").exception(E)
            wb = xl.Workbook()

        # Create sheets if they don't exist
        sheets = ["ad_svc.trace", "ad.trace", "connection.trace", "files_transferred_trace.txt"]
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)

        # Process ad_svc.trace data
        self.__write_trace_sheet_helper(wb,"ad_svc.trace", ad_svc_result)
        self.__write_trace_sheet_helper(wb, "ad.trace", ad_result)

        # Process connection.trace data
        if connection_result:
            sheet = wb["connection.trace"]

            # Set up column headers and widths
            columns = ["Connection type", "Timestamp", "Authentication", "User", "AnyDesk Connection ID"]
            widths = [15.57, 17.43, 23.14, 21, 23.43]
            Excel.__setup_worksheet(sheet, columns, widths)

            Excel.__write_connection_trace(sheet, connection_result)

        if file_trace:
            sheet = wb["files_transferred_trace.txt"]
            columns = self.rat.file_trace_keys
            widths = [15*len(columns)]
            Excel.__setup_worksheet(sheet, columns, widths)

            Excel.write_file_trace(sheet, file_trace)

        # Save workbook:
        try:
            del wb['Sheet']
        except Exception:
            pass
        wb.save(pjoin(outdir, "RatFinderResults_Anydesk.xlsx"))
        wb.close()


    @staticmethod
    def __write_teamviewer_connections(sheet, connections):
        """
        Helper function that writes the Teamviewer connections to the worksheet.
        Arguments
        ---------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to write the data to.
            connections : list[dict[str, str]]
                The list of Teamviewer connections.
        """
        for i, connection in enumerate(connections, start=2):
            for j, value in enumerate(connection.values()):
                cell_ref = f'{chr(65 + j)}{i}'
                sheet[cell_ref] = str(value)
                sheet[cell_ref].alignment = Alignment(wrap_text=True)

    def __write_teamviewer_connections_helper(self, wb, sheet_name):
        """
        Helper function used to setup the worksheet and write the Teamviewer connections.
        Arguments
        ---------
            wb : openpyxl.Workbook
                The workbook to write the data to.
            sheet_name : str
                The name of the sheet to write the data to.
        """
        sheet = wb[sheet_name]
        widths = [14.29, 10, 18, 18, 19, 19, 43, 19, 100]
        columns = self.rat.connection_file_keys
        Excel.__setup_worksheet(sheet, columns, widths)
        Excel.__write_teamviewer_connections(sheet, self.rat.connections)

    @staticmethod
    def __write_teamviewer_logs(sheet, log_results):
        """
        Helper function that writes the Teamviewer logs to the worksheet.
        Arguments
        ---------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to write the data to.
            log_results : dict[str, list[dict[str, str | dict[str, str]]]]
                The dictionary containing the Teamviewer logs.
        """
        i = 2
        for filename in log_results.keys():
            rows = log_results[filename]
            for row in rows:
                general_data = [row[key].rstrip().lstrip() for key in row.keys() if key != "Data"]
                for data_row in row["Data"]:
                    toprint = []
                    toprint.extend([str(temp).lstrip().rstrip() for temp in data_row.values()])
                    toprint.extend(general_data)
                    toprint.append(filename)
                    for j, toinsert in enumerate(toprint):
                        cell_ref = f'{chr(65 + j)}{i}'
                        sheet[cell_ref] = str(toinsert)
                        if j == 5 and toinsert:
                            sheet[cell_ref].fill = PatternFill(start_color="FF0000", end_color="FF0000",fill_type="solid")
                        sheet[cell_ref].alignment = Alignment(wrap_text=True)
                    i += 1
    def __write_teamviewer_logs_helper(self, wb, sheet_name):
        """
        Helper function used to setup the worksheet and write the Teamviewer logs.
        Arguments
        ---------
            wb : openpyxl.Workbook
                The workbook to write the data to.
            sheet_name : str
                The name of the sheet to write the data to.
        """
        sheet = wb[sheet_name]
        widths = [21, 5.29, 5.29, 8, 70, 45, 12.85, 24, 34, 9.27, 60]
        columns = self.rat.log_reporting_fields
        Excel.__setup_worksheet(sheet, columns, widths)
        Excel.__write_teamviewer_logs(sheet, self.rat.log_results)

    @staticmethod
    def __write_teamviewer_rollout(sheet, log_results):
        """
        Helper function that writes the rollout information to the worksheet.
        Arguments
        ---------
            sheet : openpyxl.worksheet.worksheet.Worksheet
                The worksheet to write the data to.
            log_results : list[dict[str, str]]
                Rollout information list
        """
        for i, entry in enumerate(log_results, start=2):
            for j, value in enumerate(entry.values()):
                cell_ref = f'{chr(65 + j)}{i}'
                sheet[cell_ref] = value
                sheet[cell_ref].alignment = Alignment(wrap_text=True)


    def __write_teamviewer_rollout_helper(self, wb, sheet_name):
        """
        Helper function used to setup the worksheet and write the rollout information.
        Arguments
        ---------
            wb : openpyxl.Workbook
                The workbook to write the data to.
            sheet_name : str
                The name of the sheet to write the data to.
        """
        sheet = wb[sheet_name]
        widths = [30, 37, 37]
        columns = self.rat.rollout_info[0].keys()
        Excel.__setup_worksheet(sheet, columns, widths)
        Excel.__write_teamviewer_rollout(sheet, self.rat.rollout_info)

    def write_teamviewer(self):
        """
        Base method used to write the Teamviewer results to an excel file.
        It generates three sheets: Connections, LogFiles and RolloutFileInformation.
        """

        outdir = pjoin(self.rat.shared.output, 'excel')
        excel_name = "TeamViewer_Results.xlsx"
        try:
            wb = xl.load_workbook(pjoin(outdir, excel_name))
        except FileNotFoundError:
            wb = xl.Workbook()
        except Exception as E:
            self.rat.shared.logger.bind(category="teamviewer").exception(E)
            wb = xl.Workbook()

        if self.rat.connections:
            wb.create_sheet("Connections")
            self.__write_teamviewer_connections_helper(wb, "Connections")

        if self.rat.log_results:
            wb.create_sheet("LogFiles")
            self.__write_teamviewer_logs_helper(wb, "LogFiles")

        if self.rat.rollout_info:
            wb.create_sheet("RolloutFileInformation")
            self.__write_teamviewer_rollout_helper(wb, "RolloutFileInformation")


        try:
            del wb['Sheet']
        except Exception:
            pass
        wb.save(pjoin(outdir, excel_name))
        wb.close()